"""
Migration worker for ThoughtSpot -> Power BI migration.

This worker:
- Reads uploaded .tml, .yaml, .yml, .json, .csv, .xlsx, .xls, and .zip files
- Extracts ThoughtSpot-style metadata
- Builds workbooks, worksheets, calculated fields, tables, relationships, and DAX conversions
- Saves a frontend-compatible result structure
"""

from typing import List, Any, Dict, Optional
from datetime import datetime
from pathlib import Path
import json
import csv
import zipfile
import tempfile
import re

from loguru import logger

from api.models.api_models import JobStatus
from storage.job_store import JobStore
from storage.result_store import ResultStore


# ============================================================
# File helpers
# ============================================================

def _detect_file_type(file_path: str) -> str:
    suffix = Path(file_path).suffix.lower()

    if suffix in [".tml", ".yaml", ".yml"]:
        return "thoughtspot_tml"

    if suffix == ".json":
        return "json"

    if suffix == ".zip":
        return "zip"

    if suffix == ".csv":
        return "csv"

    if suffix in [".xlsx", ".xls"]:
        return "excel"

    return "unknown"


def _build_file_summary(file_paths: List[str]) -> List[Dict[str, Any]]:
    files = []

    for file_path in file_paths:
        path = Path(file_path)

        files.append(
            {
                "filename": path.name,
                "file_path": str(path),
                "file_type": _detect_file_type(str(path)),
                "exists": path.exists(),
                "size_bytes": path.stat().st_size if path.exists() else 0,
            }
        )

    return files


def _extract_zip_files(file_path: str) -> List[str]:
    extracted_files = []

    if not zipfile.is_zipfile(file_path):
        return extracted_files

    temp_dir = Path(tempfile.mkdtemp(prefix="thoughtspot_zip_"))

    with zipfile.ZipFile(file_path, "r") as zip_ref:
        zip_ref.extractall(temp_dir)

    for path in temp_dir.rglob("*"):
        if path.is_file():
            extracted_files.append(str(path))

    return extracted_files


def _safe_read_text(file_path: str) -> str:
    path = Path(file_path)

    try:
        return path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        return path.read_text(encoding="latin-1", errors="ignore")


def _safe_name(value: Any, fallback: str) -> str:
    if value is None:
        return fallback

    text = str(value).strip().strip('"').strip("'")

    return text if text else fallback


# ============================================================
# Data type detection
# ============================================================

def _is_int(value: Any) -> bool:
    try:
        text = str(value).strip()
        if text == "":
            return False
        int(text)
        return True
    except Exception:
        return False


def _is_float(value: Any) -> bool:
    try:
        text = str(value).strip()
        if text == "":
            return False
        float(text)
        return True
    except Exception:
        return False


def _is_date(value: Any) -> bool:
    text = str(value).strip()

    if not text:
        return False

    date_patterns = [
        r"^\d{4}-\d{2}-\d{2}$",
        r"^\d{2}-\d{2}-\d{4}$",
        r"^\d{2}/\d{2}/\d{4}$",
        r"^\d{4}/\d{2}/\d{2}$",
    ]

    return any(re.match(pattern, text) for pattern in date_patterns)


def _detect_column_type(column_name: str, values: List[Any]) -> str:
    clean_name = column_name.lower()

    if any(token in clean_name for token in ["date", "created_at", "updated_at", "time"]):
        return "date"

    if clean_name.endswith("id") or clean_name.endswith("_id") or " id" in clean_name:
        return "string"

    sample_values = [
        value for value in values
        if value is not None and str(value).strip() != ""
    ][:50]

    if not sample_values:
        return "string"

    if all(_is_date(value) for value in sample_values):
        return "date"

    if all(_is_int(value) for value in sample_values):
        return "integer"

    if all(_is_float(value) for value in sample_values):
        return "double"

    return "string"


def _normalize_datatype(data_type: str) -> str:
    text = str(data_type or "").lower()

    if text in ["int", "integer", "bigint", "smallint", "number"]:
        return "integer"

    if text in ["float", "double", "decimal", "numeric", "currency"]:
        return "double"

    if text in ["date", "datetime", "timestamp"]:
        return "date"

    if text in ["bool", "boolean"]:
        return "boolean"

    return "string"


# ============================================================
# CSV / Excel parsing
# ============================================================

def _parse_csv_table(file_path: str) -> Dict[str, Any]:
    path = Path(file_path)

    rows = []
    columns = []

    with open(path, "r", encoding="utf-8", errors="ignore", newline="") as file:
        reader = csv.DictReader(file)
        columns = reader.fieldnames or []

        for index, row in enumerate(reader):
            if index < 200:
                rows.append(row)
            else:
                break

    total_rows = 0

    with open(path, "r", encoding="utf-8", errors="ignore", newline="") as file:
        reader = csv.reader(file)
        total_rows = max(sum(1 for _ in reader) - 1, 0)

    column_details = []

    for column in columns:
        sample_values = [row.get(column) for row in rows]
        detected_type = _detect_column_type(column, sample_values)

        column_details.append(
            {
                "name": column,
                "display_name": column,
                "data_type": detected_type,
                "datatype": detected_type,
                "role": "attribute" if detected_type in ["string", "date"] else "measure",
            }
        )

    return {
        "table_name": path.stem,
        "display_name": path.stem,
        "row_count": total_rows,
        "column_details": column_details,
        "columns": column_details,
        "data_preview": rows[:20],
    }


def _parse_excel_table(file_path: str) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception:
        logger.warning("openpyxl is not installed. Excel parsing skipped.")
        return []

    path = Path(file_path)
    workbook = load_workbook(path, read_only=True, data_only=True)

    tables = []

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            continue

        headers = [
            str(value).strip() if value is not None else f"Column_{index + 1}"
            for index, value in enumerate(rows[0])
        ]

        preview_rows = []

        for row in rows[1:201]:
            preview_rows.append(
                {
                    headers[index]: value
                    for index, value in enumerate(row)
                    if index < len(headers)
                }
            )

        column_details = []

        for column in headers:
            sample_values = [row.get(column) for row in preview_rows]
            detected_type = _detect_column_type(column, sample_values)

            column_details.append(
                {
                    "name": column,
                    "display_name": column,
                    "data_type": detected_type,
                    "datatype": detected_type,
                    "role": "attribute" if detected_type in ["string", "date"] else "measure",
                }
            )

        tables.append(
            {
                "table_name": sheet.title,
                "display_name": sheet.title,
                "row_count": max(len(rows) - 1, 0),
                "column_details": column_details,
                "columns": column_details,
                "data_preview": preview_rows[:20],
            }
        )

    return tables


# ============================================================
# TML parsing
# ============================================================

def _extract_yaml_value(text: str, key: str, default: str = "") -> str:
    pattern = rf"^\s*{re.escape(key)}\s*:\s*(.+?)\s*$"
    match = re.search(pattern, text, flags=re.MULTILINE)

    if not match:
        return default

    return match.group(1).strip().strip('"').strip("'")


def _extract_list_names_after_key(text: str, key: str) -> List[str]:
    """
    Lightweight YAML/TML list extractor.
    Example:
    worksheets:
      - name: Sales
      - name: Profit
    """
    results = []

    key_match = re.search(
        rf"^\s*{re.escape(key)}\s*:\s*$",
        text,
        flags=re.IGNORECASE | re.MULTILINE,
    )

    if not key_match:
        return results

    block_start = key_match.end()
    next_top_level = re.search(r"^\S.*:\s*$", text[block_start:], flags=re.MULTILINE)

    if next_top_level:
        block = text[block_start:block_start + next_top_level.start()]
    else:
        block = text[block_start:]

    name_matches = re.findall(
        r"^\s*-\s+name\s*:\s*(.+?)\s*$",
        block,
        flags=re.IGNORECASE | re.MULTILINE,
    )

    for name in name_matches:
        clean_name = name.strip().strip('"').strip("'")
        if clean_name:
            results.append(clean_name)

    return results


def _parse_tml_file(file_path: str) -> Dict[str, Any]:
    path = Path(file_path)
    text = _safe_read_text(file_path)

    object_name = (
        _extract_yaml_value(text, "name")
        or _extract_yaml_value(text, "display_name")
        or path.stem
    )

    formulas = []

    formula_blocks = re.findall(
        r"-\s+name\s*:\s*(.+?)\n\s+(?:expr|formula|calc_formula)\s*:\s*(.+?)(?:\n|$)",
        text,
        flags=re.IGNORECASE,
    )

    for index, (name, expr) in enumerate(formula_blocks):
        clean_name = _safe_name(name, f"Calculated Field {index + 1}")
        clean_expr = _safe_name(expr, "")

        formulas.append(
            {
                "id": f"calc_{path.stem}_{index + 1}",
                "calc_id": f"calc_{path.stem}_{index + 1}",
                "name": clean_name,
                "caption": clean_name,
                "formula": clean_expr,
                "calc_formula": clean_expr,
                "role": "measure",
                "calc_type": "measure",
                "datatype": "double",
                "data_type": "double",
                "source_file": path.name,
            }
        )

    if not formulas:
        expr_lines = re.findall(
            r"^\s*(?:expr|formula|calc_formula)\s*:\s*(.+)",
            text,
            flags=re.IGNORECASE | re.MULTILINE,
        )

        for index, expr in enumerate(expr_lines):
            clean_expr = _safe_name(expr, "")
            clean_name = f"Calculated Field {index + 1}"

            formulas.append(
                {
                    "id": f"calc_{path.stem}_{index + 1}",
                    "calc_id": f"calc_{path.stem}_{index + 1}",
                    "name": clean_name,
                    "caption": clean_name,
                    "formula": clean_expr,
                    "calc_formula": clean_expr,
                    "role": "measure",
                    "calc_type": "measure",
                    "datatype": "double",
                    "data_type": "double",
                    "source_file": path.name,
                }
            )

    worksheet_names = _extract_list_names_after_key(text, "worksheets")

    if not worksheet_names:
        worksheet_names = _extract_list_names_after_key(text, "answers")

    formula_names = {field["name"] for field in formulas}
    worksheets = []

    for index, worksheet_name in enumerate(worksheet_names):
        if worksheet_name in formula_names:
            continue

        worksheets.append(
            {
                "id": f"worksheet_{path.stem}_{index + 1}",
                "name": worksheet_name,
                "chart_type": "bar",
                "visual_type": "bar",
                "dimensions": ["Region"] if re.search(r"\bRegion\b", text) else [],
                "measures": [
                    {
                        "name": formulas[0]["name"],
                        "type": "calculated",
                    }
                ] if formulas else [],
                "filters": [],
                "axes": {
                    "rows": "Region" if re.search(r"\bRegion\b", text) else "",
                    "columns": formulas[0]["name"] if formulas else "",
                },
                "source_file": path.name,
            }
        )

    if not worksheets:
        worksheets.append(
            {
                "id": f"worksheet_{path.stem}_1",
                "name": f"{object_name} Worksheet",
                "chart_type": "bar",
                "visual_type": "bar",
                "dimensions": [],
                "measures": [
                    {
                        "name": formula["name"],
                        "type": "calculated",
                    }
                    for formula in formulas[:3]
                ],
                "filters": [],
                "axes": {
                    "rows": "",
                    "columns": formulas[0]["name"] if formulas else "",
                },
                "source_file": path.name,
            }
        )

    # Basic join extraction from TML
    joins = []

    join_blocks = re.findall(
        r"source_table\s*:\s*(.+?)\n\s+target_table\s*:\s*(.+?)\n\s+source_key\s*:\s*(.+?)\n\s+target_key\s*:\s*(.+?)(?:\n|$)",
        text,
        flags=re.IGNORECASE,
    )

    for index, (source_table, target_table, source_key, target_key) in enumerate(join_blocks):
        joins.append(
            {
                "relationship_id": f"tml_rel_{path.stem}_{index + 1}",
                "source_table": _safe_name(source_table, ""),
                "target_table": _safe_name(target_table, ""),
                "source_column": _safe_name(source_key, ""),
                "target_column": _safe_name(target_key, ""),
                "relationship_type": "many_to_one",
                "confidence_score": 0.95,
                "detection_method": "tml_join_metadata",
            }
        )

    return {
        "filename": path.name,
        "name": object_name,
        "object_type": "liveboard",
        "worksheets": worksheets,
        "calculated_fields": formulas,
        "data_sources": [],
        "relationships": joins,
    }


# ============================================================
# JSON parsing
# ============================================================

def _normalize_calculated_field(field: Any, index: int, source_prefix: str = "json") -> Dict[str, Any]:
    if isinstance(field, dict):
        name = (
            field.get("name")
            or field.get("caption")
            or field.get("display_name")
            or f"Calculated Field {index + 1}"
        )

        formula = (
            field.get("formula")
            or field.get("expr")
            or field.get("calc_formula")
            or field.get("source_formula")
            or ""
        )

        calc_id = (
            field.get("calc_id")
            or field.get("id")
            or f"calc_{source_prefix}_{index + 1}"
        )

        calc_type = (
            field.get("calc_type")
            or field.get("role")
            or field.get("type")
            or "measure"
        )

        datatype = _normalize_datatype(
            field.get("datatype")
            or field.get("data_type")
            or field.get("type")
            or "double"
        )

        return {
            "id": str(calc_id),
            "calc_id": str(calc_id),
            "name": str(name),
            "caption": str(field.get("caption") or name),
            "formula": str(formula),
            "calc_formula": str(formula),
            "role": calc_type,
            "calc_type": calc_type,
            "datatype": datatype,
            "data_type": datatype,
        }

    return {
        "id": f"calc_{source_prefix}_{index + 1}",
        "calc_id": f"calc_{source_prefix}_{index + 1}",
        "name": f"Calculated Field {index + 1}",
        "caption": f"Calculated Field {index + 1}",
        "formula": str(field),
        "calc_formula": str(field),
        "role": "measure",
        "calc_type": "measure",
        "datatype": "double",
        "data_type": "double",
    }


def _normalize_worksheet(worksheet: Any, index: int, source_prefix: str = "json") -> Dict[str, Any]:
    if isinstance(worksheet, dict):
        name = (
            worksheet.get("name")
            or worksheet.get("title")
            or worksheet.get("display_name")
            or f"Worksheet {index + 1}"
        )

        return {
            "id": worksheet.get("id") or f"worksheet_{source_prefix}_{index + 1}",
            "name": name,
            "chart_type": worksheet.get("chart_type") or worksheet.get("visual_type") or "table",
            "visual_type": worksheet.get("visual_type") or worksheet.get("chart_type") or "table",
            "dimensions": worksheet.get("dimensions") or [],
            "measures": worksheet.get("measures") or [],
            "filters": worksheet.get("filters") or [],
            "axes": worksheet.get("axes") or {},
        }

    return {
        "id": f"worksheet_{source_prefix}_{index + 1}",
        "name": str(worksheet),
        "chart_type": "table",
        "visual_type": "table",
        "dimensions": [],
        "measures": [],
        "filters": [],
        "axes": {},
    }


def _normalize_table(table: Dict[str, Any], index: int) -> Dict[str, Any]:
    table_name = (
        table.get("table_name")
        or table.get("name")
        or table.get("display_name")
        or f"Table_{index + 1}"
    )

    raw_columns = (
        table.get("column_details")
        or table.get("columns")
        or table.get("fields")
        or []
    )

    column_details = []

    for col_index, column in enumerate(raw_columns):
        if isinstance(column, dict):
            column_name = (
                column.get("name")
                or column.get("column_name")
                or column.get("display_name")
                or f"Column_{col_index + 1}"
            )

            datatype = _normalize_datatype(
                column.get("data_type")
                or column.get("datatype")
                or column.get("type")
                or "string"
            )
        else:
            column_name = str(column)
            datatype = "string"

        column_details.append(
            {
                "name": column_name,
                "display_name": column_name,
                "data_type": datatype,
                "datatype": datatype,
                "role": "attribute" if datatype in ["string", "date"] else "measure",
            }
        )

    return {
        "table_name": table_name,
        "display_name": table.get("display_name") or table_name,
        "row_count": table.get("row_count") or 0,
        "column_details": column_details,
        "columns": column_details,
        "data_preview": table.get("data_preview") or [],
    }


def _parse_json_file(file_path: str) -> Dict[str, Any]:
    path = Path(file_path)

    try:
        data = json.loads(_safe_read_text(file_path))
    except Exception as error:
        logger.warning(f"Could not parse JSON file {file_path}: {error}")
        return {
            "filename": path.name,
            "name": path.stem,
            "object_type": "json",
            "worksheets": [],
            "calculated_fields": [],
            "data_sources": [],
            "relationships": [],
        }

    objects = data.get("objects") or data.get("workbooks") or []

    all_worksheets = []
    all_calculated_fields = []
    all_data_sources = []
    all_relationships = []

    # IMPORTANT FIX:
    # Parse all JSON objects, not only first object.
    if isinstance(objects, list) and objects:
        for obj_index, obj in enumerate(objects):
            if not isinstance(obj, dict):
                continue

            prefix = f"{path.stem}_{obj_index + 1}"

            worksheets = obj.get("worksheets") or obj.get("answers") or []
            calculated_fields = obj.get("calculated_fields") or obj.get("formulas") or []
            data_sources = obj.get("data_sources") or obj.get("sources") or []
            relationships = obj.get("relationships") or obj.get("joins") or []

            all_worksheets.extend(
                _normalize_worksheet(worksheet, index, prefix)
                for index, worksheet in enumerate(worksheets)
            )

            all_calculated_fields.extend(
                _normalize_calculated_field(field, index, prefix)
                for index, field in enumerate(calculated_fields)
            )

            normalized_sources = []

            for source in data_sources:
                if not isinstance(source, dict):
                    continue

                tables = source.get("table_details") or source.get("tables") or []

                normalized_tables = [
                    _normalize_table(table, table_index)
                    for table_index, table in enumerate(tables)
                    if isinstance(table, dict)
                ]

                normalized_sources.append(
                    {
                        "name": source.get("name") or source.get("display_name") or "JSON Data Source",
                        "table_details": normalized_tables,
                    }
                )

            all_data_sources.extend(normalized_sources)

            for rel_index, relationship in enumerate(relationships):
                if isinstance(relationship, dict):
                    all_relationships.append(
                        {
                            "relationship_id": relationship.get("relationship_id") or f"json_rel_{prefix}_{rel_index + 1}",
                            "source_table": relationship.get("source_table") or relationship.get("from_table") or "",
                            "target_table": relationship.get("target_table") or relationship.get("to_table") or "",
                            "source_column": relationship.get("source_column") or relationship.get("from_column") or relationship.get("source_key") or "",
                            "target_column": relationship.get("target_column") or relationship.get("to_column") or relationship.get("target_key") or "",
                            "relationship_type": relationship.get("relationship_type") or "many_to_one",
                            "confidence_score": relationship.get("confidence_score") or 0.9,
                            "detection_method": "json_metadata",
                        }
                    )

        return {
            "filename": path.name,
            "name": path.stem,
            "object_type": "json_metadata",
            "worksheets": all_worksheets,
            "calculated_fields": all_calculated_fields,
            "data_sources": all_data_sources,
            "relationships": all_relationships,
        }

    calculated_fields = data.get("calculated_fields") or data.get("formulas") or []
    worksheets = data.get("worksheets") or data.get("answers") or []
    data_sources = data.get("data_sources") or []

    return {
        "filename": path.name,
        "name": data.get("name") or path.stem,
        "object_type": data.get("object_type") or "json_metadata",
        "worksheets": [
            _normalize_worksheet(worksheet, index, path.stem)
            for index, worksheet in enumerate(worksheets)
        ],
        "calculated_fields": [
            _normalize_calculated_field(field, index, path.stem)
            for index, field in enumerate(calculated_fields)
        ],
        "data_sources": data_sources,
        "relationships": data.get("relationships") or [],
    }


# ============================================================
# Relationship detection
# ============================================================

def _column_names(table: Dict[str, Any]) -> List[str]:
    columns = table.get("column_details") or table.get("columns") or []

    names = []

    for column in columns:
        if isinstance(column, dict):
            name = column.get("name") or column.get("display_name")
            if name:
                names.append(str(name))
        else:
            names.append(str(column))

    return names


def _normalize_key_name(name: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(name).lower())


def _detect_relationships(tables: List[Dict[str, Any]], existing_relationships: Optional[List[Dict[str, Any]]] = None) -> List[Dict[str, Any]]:
    relationships = []
    seen = set()

    existing_relationships = existing_relationships or []

    for index, rel in enumerate(existing_relationships):
        source_table = rel.get("source_table") or rel.get("from_table") or ""
        target_table = rel.get("target_table") or rel.get("to_table") or ""
        source_column = rel.get("source_column") or rel.get("source_key") or rel.get("from_column") or ""
        target_column = rel.get("target_column") or rel.get("target_key") or rel.get("to_column") or ""

        if not source_table or not target_table or not source_column or not target_column:
            continue

        key = (
            _normalize_key_name(source_table),
            _normalize_key_name(source_column),
            _normalize_key_name(target_table),
            _normalize_key_name(target_column),
        )

        if key in seen:
            continue

        seen.add(key)

        relationships.append(
            {
                "relationship_id": rel.get("relationship_id") or f"rel_existing_{index + 1}",
                "source_table": source_table,
                "target_table": target_table,
                "source_column": source_column,
                "target_column": target_column,
                "relationship_type": rel.get("relationship_type") or "many_to_one",
                "cardinality": rel.get("cardinality") or "many_to_one",
                "confidence_score": rel.get("confidence_score") or 0.95,
                "detection_method": rel.get("detection_method") or "source_metadata",
                "active": True,
            }
        )

    # Detect relationships from matching ID columns
    for left_index, left_table in enumerate(tables):
        left_table_name = left_table.get("table_name") or left_table.get("display_name") or f"Table_{left_index + 1}"
        left_columns = _column_names(left_table)

        for right_index, right_table in enumerate(tables):
            if left_index >= right_index:
                continue

            right_table_name = right_table.get("table_name") or right_table.get("display_name") or f"Table_{right_index + 1}"
            right_columns = _column_names(right_table)

            for left_column in left_columns:
                left_norm = _normalize_key_name(left_column)

                if not (
                    left_norm.endswith("id")
                    or left_norm in ["id", "key"]
                    or "customerid" in left_norm
                    or "productid" in left_norm
                    or "orderid" in left_norm
                ):
                    continue

                for right_column in right_columns:
                    right_norm = _normalize_key_name(right_column)

                    if left_norm != right_norm:
                        continue

                    key = (
                        _normalize_key_name(left_table_name),
                        left_norm,
                        _normalize_key_name(right_table_name),
                        right_norm,
                    )

                    reverse_key = (
                        _normalize_key_name(right_table_name),
                        right_norm,
                        _normalize_key_name(left_table_name),
                        left_norm,
                    )

                    if key in seen or reverse_key in seen:
                        continue

                    seen.add(key)

                    relationships.append(
                        {
                            "relationship_id": f"rel_{len(relationships) + 1}",
                            "source_table": left_table_name,
                            "target_table": right_table_name,
                            "source_column": left_column,
                            "target_column": right_column,
                            "relationship_type": "many_to_one",
                            "cardinality": "many_to_one",
                            "confidence_score": 0.85,
                            "detection_method": "matching_key_columns",
                            "active": True,
                        }
                    )

    return relationships


# ============================================================
# DAX conversion
# ============================================================

def _replace_operators(formula: str) -> str:
    """
    Convert common ThoughtSpot/operator syntax into DAX-compatible syntax.
    """
    dax = str(formula or "")

    # ThoughtSpot / SQL style not equal operators
    dax = re.sub(r"(?<![<>=!])!=(?!=)", "<>", dax)

    # Boolean operators
    dax = re.sub(r"\band\b", "&&", dax, flags=re.IGNORECASE)
    dax = re.sub(r"\bor\b", "||", dax, flags=re.IGNORECASE)
    dax = re.sub(r"\bnot\b", "NOT", dax, flags=re.IGNORECASE)

    return dax


def _replace_basic_functions(formula: str) -> str:
    """
    Convert common aggregate and scalar ThoughtSpot functions to DAX names.
    """
    dax = str(formula or "")

    function_replacements = {
        "count_distinct": "DISTINCTCOUNT",
        "countdistinct": "DISTINCTCOUNT",
        "average": "AVERAGE",
        "avg": "AVERAGE",
        "sum": "SUM",
        "count": "COUNT",
        "min": "MIN",
        "max": "MAX",
        "if": "IF",
        "coalesce": "COALESCE",
        "year": "YEAR",
        "month": "MONTH",
        "day": "DAY",
        "hour": "HOUR",
        "minute": "MINUTE",
        "second": "SECOND",
        "today": "TODAY",
        "now": "NOW",
        "upper": "UPPER",
        "lower": "LOWER",
        "left": "LEFT",
        "right": "RIGHT",
        "len": "LEN",
        "length": "LEN",
        "trim": "TRIM",
        "abs": "ABS",
        "round": "ROUND",
        "floor": "FLOOR",
        "ceiling": "CEILING",
        "sqrt": "SQRT",
    }

    # Replace longest names first so count_distinct is handled before count.
    for source_function in sorted(function_replacements.keys(), key=len, reverse=True):
        target_function = function_replacements[source_function]
        pattern = rf"\b{re.escape(source_function)}\s*\("
        dax = re.sub(pattern, f"{target_function}(", dax, flags=re.IGNORECASE)

    return dax


def _replace_null_functions(formula: str) -> str:
    """
    Convert common null checks into DAX style.
    """
    dax = str(formula or "")

    # ifnull(x, y) -> COALESCE(x, y)
    dax = re.sub(
        r"\bifnull\s*\(",
        "COALESCE(",
        dax,
        flags=re.IGNORECASE,
    )

    # isnull(x) -> ISBLANK(x)
    dax = re.sub(
        r"\bisnull\s*\(",
        "ISBLANK(",
        dax,
        flags=re.IGNORECASE,
    )

    # notnull(x) -> NOT(ISBLANK(x)) for simple single-argument usage.
    dax = re.sub(
        r"\bnotnull\s*\(([^()]+)\)",
        r"NOT(ISBLANK(\1))",
        dax,
        flags=re.IGNORECASE,
    )

    return dax


def _replace_text_functions(formula: str) -> str:
    """
    Convert common text functions.
    """
    dax = str(formula or "")

    # contains([Name], "abc") -> CONTAINSSTRING([Name], "abc")
    dax = re.sub(
        r"\bcontains\s*\(",
        "CONTAINSSTRING(",
        dax,
        flags=re.IGNORECASE,
    )

    # concat(a,b) -> CONCATENATE(a,b)
    dax = re.sub(
        r"\bconcat\s*\(",
        "CONCATENATE(",
        dax,
        flags=re.IGNORECASE,
    )

    return dax


def _replace_date_functions(formula: str) -> str:
    """
    Convert common date functions.
    """
    dax = str(formula or "")

    # date_diff / datediff becomes DATEDIFF. Argument order may still need review.
    dax = re.sub(
        r"\bdate_diff\s*\(",
        "DATEDIFF(",
        dax,
        flags=re.IGNORECASE,
    )

    dax = re.sub(
        r"\bdatediff\s*\(",
        "DATEDIFF(",
        dax,
        flags=re.IGNORECASE,
    )

    return dax


def _convert_simple_case_when(formula: str) -> Optional[str]:
    """
    Convert a simple CASE WHEN formula to SWITCH(TRUE()).

    Supported pattern:
    case
      when condition then value
      when condition then value
      else value
    end

    This does not handle deeply nested CASE expressions.
    """
    text = str(formula or "").strip()

    if not re.search(r"\bcase\b", text, flags=re.IGNORECASE):
        return None

    if not re.search(r"\bwhen\b", text, flags=re.IGNORECASE):
        return None

    when_then_pairs = re.findall(
        r"\bwhen\s+(.+?)\s+then\s+(.+?)(?=\s+when\s+|\s+else\s+|\s+end\b)",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )

    else_match = re.search(
        r"\belse\s+(.+?)\s+end\b",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )

    if not when_then_pairs:
        return None

    parts = ["SWITCH(TRUE()"]

    for condition, value in when_then_pairs:
        clean_condition = _replace_operators(condition.strip())
        clean_value = value.strip()
        parts.append(f", {clean_condition}, {clean_value}")

    if else_match:
        parts.append(f", {else_match.group(1).strip()}")
    else:
        parts.append(", BLANK()")

    parts.append(")")

    return "".join(parts)


def _thoughtspot_formula_to_dax(formula: str) -> str:
    """
    Convert ThoughtSpot-style formulas into Power BI DAX-style formulas.

    This converter handles common formulas automatically and leaves complex
    formulas available for manual review through validation warnings.
    """
    if not formula:
        return ""

    original_formula = str(formula).strip()

    # First handle simple CASE WHEN because it needs structural conversion.
    case_dax = _convert_simple_case_when(original_formula)
    if case_dax:
        dax = case_dax
    else:
        dax = original_formula

    dax = _replace_null_functions(dax)
    dax = _replace_text_functions(dax)
    dax = _replace_date_functions(dax)
    dax = _replace_basic_functions(dax)
    dax = _replace_operators(dax)

    # Clean extra whitespace
    dax = re.sub(r"\s+", " ", dax).strip()

    return dax


def _validate_dax_conversion(source_formula: str, dax_formula: str) -> Dict[str, Any]:
    """
    Validate conversion quality and decide whether it is auto-validated
    or needs manual review.
    """
    warnings = []

    source = str(source_formula or "")
    lower_formula = source.lower()

    if not dax_formula:
        return {
            "status": "failed",
            "confidence_score": 0.0,
            "warnings": ["Empty DAX formula"],
        }

    if dax_formula.count("(") != dax_formula.count(")"):
        return {
            "status": "manual_review",
            "confidence_score": 0.45,
            "warnings": ["Parentheses may be unbalanced"],
        }

    complex_patterns = {
        "running_sum": "Running totals usually need CALCULATE with filter context.",
        "moving_average": "Moving averages usually need DATESINPERIOD or window logic.",
        "rank(": "Rank logic should be reviewed and converted with RANKX.",
        "percentile": "Percentile logic should be reviewed and converted with PERCENTILEX functions.",
        "window_": "Window functions need manual DAX review.",
        "lag(": "Lag logic usually needs DATEADD, OFFSET, or custom row context logic.",
        "lead(": "Lead logic usually needs OFFSET or custom row context logic.",
    }

    for pattern, message in complex_patterns.items():
        if pattern in lower_formula:
            warnings.append(message)

    # CASE WHEN is supported only for simple patterns.
    if "case" in lower_formula and "SWITCH(TRUE()" not in dax_formula:
        warnings.append("CASE WHEN formula is complex and needs manual review.")

    # DATEDIFF argument order may need review depending on ThoughtSpot syntax.
    if "date_diff" in lower_formula or "datediff" in lower_formula:
        warnings.append("DATEDIFF conversion may need date-part/order review.")

    if warnings:
        return {
            "status": "manual_review",
            "confidence_score": 0.70,
            "warnings": warnings,
        }

    # If formula changed and no warning, mark as high confidence.
    if source.strip() != dax_formula.strip():
        return {
            "status": "validated",
            "confidence_score": 0.92,
            "warnings": [],
        }

    # If formula did not change, it might already be DAX or might need review.
    return {
        "status": "manual_review",
        "confidence_score": 0.65,
        "warnings": ["Formula was not changed. Please review if it is already valid DAX."],
    }


def _build_conversions(calculated_fields: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    conversions = []

    for index, field in enumerate(calculated_fields):
        calc_id = field.get("calc_id") or field.get("id") or f"calc_{index + 1}"
        formula = field.get("calc_formula") or field.get("formula") or ""
        name = field.get("name") or field.get("caption") or f"Calculation {index + 1}"

        dax_formula = _thoughtspot_formula_to_dax(formula)
        validation = _validate_dax_conversion(formula, dax_formula)

        conversion_method = (
            "RULE_BASED_ADVANCED"
            if validation["status"] == "validated"
            else "RULE_BASED_REVIEW_REQUIRED"
        )

        conversions.append(
            {
                "conversion_id": f"conv_{calc_id}",
                "calc_id": calc_id,
                "source_calculated_field": name,
                "source_name": name,
                "source_formula": formula,
                "dax_formula": dax_formula,
                "converted_dax_formula": dax_formula,
                "conversion_method": conversion_method,
                "confidence_score": validation["confidence_score"],
                "status": validation["status"],
                "warnings": validation["warnings"],
            }
        )

    return conversions


# ============================================================
# Main parser
# ============================================================

def _parse_uploaded_files(file_paths: List[str]) -> Dict[str, Any]:
    all_paths = []

    for file_path in file_paths:
        if _detect_file_type(file_path) == "zip":
            all_paths.extend(_extract_zip_files(file_path))
        else:
            all_paths.append(file_path)

    workbooks = []
    standalone_tables = []
    existing_relationships = []

    for file_path in all_paths:
        file_type = _detect_file_type(file_path)

        try:
            if file_type == "thoughtspot_tml":
                workbook = _parse_tml_file(file_path)
                workbooks.append(workbook)
                existing_relationships.extend(workbook.get("relationships") or [])

            elif file_type == "json":
                workbook = _parse_json_file(file_path)
                workbooks.append(workbook)
                existing_relationships.extend(workbook.get("relationships") or [])

            elif file_type == "csv":
                standalone_tables.append(_parse_csv_table(file_path))

            elif file_type == "excel":
                standalone_tables.extend(_parse_excel_table(file_path))

        except Exception as error:
            logger.warning(f"Failed to parse {file_path}: {error}", exc_info=True)

    if standalone_tables:
        if not workbooks:
            workbooks.append(
                {
                    "filename": "source_data",
                    "name": "Source Data",
                    "object_type": "data_source",
                    "worksheets": [],
                    "calculated_fields": [],
                    "data_sources": [],
                    "relationships": [],
                }
            )

        workbooks[0].setdefault("data_sources", [])

        workbooks[0]["data_sources"].append(
            {
                "name": "Uploaded Source Data",
                "table_details": standalone_tables,
            }
        )

    all_calculated_fields = []

    for workbook in workbooks:
        all_calculated_fields.extend(workbook.get("calculated_fields") or [])

    all_tables = []

    for workbook in workbooks:
        for data_source in workbook.get("data_sources") or []:
            all_tables.extend(data_source.get("table_details") or [])

    relationships = _detect_relationships(all_tables, existing_relationships)
    conversions = _build_conversions(all_calculated_fields)

    total_worksheets = sum(
        len(workbook.get("worksheets") or []) for workbook in workbooks
    )

    validated_count = len(
        [conversion for conversion in conversions if conversion.get("status") == "validated"]
    )

    manual_review_count = len(
        [conversion for conversion in conversions if conversion.get("status") == "manual_review"]
    )

    failed_count = len(
        [conversion for conversion in conversions if conversion.get("status") == "failed"]
    )

    summary = {
        "total_dashboards": len(workbooks),
        "total_worksheets": total_worksheets,
        "total_tables": len(all_tables),
        "total_calculated_fields": len(all_calculated_fields),
        "uploaded_file_count": len(file_paths),
        "valid_file_count": len(all_paths),
        "object_count": len(workbooks),
        "formula_count": len(all_calculated_fields),
        "relationship_count": len(relationships),
        "conversion_count": len(conversions),
        "validated_conversion_count": validated_count,
        "manual_review_count": manual_review_count,
        "failed_conversion_count": failed_count,
        "message": "ThoughtSpot metadata parsed successfully.",
    }

    return {
        "workbooks": workbooks,
        "tables": all_tables,
        "calculations": all_calculated_fields,
        "formulas": all_calculated_fields,
        "conversions": conversions,
        "relationships": relationships,
        "suggested_relationships": relationships,
        "summary": summary,
    }


# ============================================================
# Worker entrypoint
# ============================================================

def execute_thoughtspot_powerbi_migration(job_id: str, file_paths: List[str]):
    job_store = JobStore()
    result_store = ResultStore()

    try:
        logger.info(f"Starting ThoughtSpot -> Power BI migration job: {job_id}")

        job_store.update_status(
            job_id=job_id,
            status=JobStatus.RUNNING,
        )

        try:
            job_store.update_progress(
                job_id=job_id,
                percent=10,
                stage="initializing",
                message="Initializing ThoughtSpot to Power BI migration",
            )
        except TypeError:
            logger.warning("JobStore.update_progress signature is different. Skipping progress update.")

        file_summary = _build_file_summary(file_paths)

        valid_files = [
            file_info for file_info in file_summary if file_info.get("exists")
        ]

        try:
            job_store.update_progress(
                job_id=job_id,
                percent=40,
                stage="parsing_files",
                message=f"Parsing {len(valid_files)} valid uploaded file(s)",
            )
        except TypeError:
            logger.warning("JobStore.update_progress signature is different. Skipping progress update.")

        parsed_data = _parse_uploaded_files(file_paths)

        try:
            job_store.update_progress(
                job_id=job_id,
                percent=75,
                stage="generating_conversions",
                message="Generating DAX conversions and relationships",
            )
        except TypeError:
            logger.warning("JobStore.update_progress signature is different. Skipping progress update.")

        result = {
            "job_id": job_id,
            "migration_id": job_id,
            "type": "thoughtspot_powerbi_migration_result",
            "status": "completed",
            "source": "thoughtspot",
            "target": "powerbi",
            "generated_at": datetime.utcnow().isoformat(),
            "summary": parsed_data["summary"],
            "files": file_summary,
            "workbooks": parsed_data["workbooks"],
            "objects": parsed_data["workbooks"],
            "tables": parsed_data["tables"],
            "calculations": parsed_data["calculations"],
            "formulas": parsed_data["formulas"],
            "conversions": parsed_data["conversions"],
            "relationships": parsed_data["relationships"],
            "suggested_relationships": parsed_data["suggested_relationships"],
            "powerbi_artifacts": {
                "model_bim": None,
                "report_json": None,
                "pbix_file": None,
            },
        }

        try:
            job_store.update_progress(
                job_id=job_id,
                percent=90,
                stage="saving_result",
                message="Saving migration result",
            )
        except TypeError:
            logger.warning("JobStore.update_progress signature is different. Skipping progress update.")

        result_file_path = result_store.save_result(
            job_id=job_id,
            result=result,
        )

        try:
            job_store.update_progress(
                job_id=job_id,
                percent=100,
                stage="completed",
                message="ThoughtSpot to Power BI migration completed successfully",
            )
        except TypeError:
            logger.warning("JobStore.update_progress signature is different. Skipping progress update.")

        try:
            job_store.update_status(
                job_id=job_id,
                status=JobStatus.COMPLETED,
                total_objects=parsed_data["summary"]["object_count"],
                formulas_converted=parsed_data["summary"]["conversion_count"],
                relationships_created=parsed_data["summary"]["relationship_count"],
                result_file_path=result_file_path,
            )
        except TypeError:
            job_store.update_status(
                job_id=job_id,
                status=JobStatus.COMPLETED,
            )

        final_result = {
            **result,
            "result_file_path": result_file_path,
        }

        logger.info(f"Migration job completed successfully: {job_id}")

        return final_result

    except Exception as error:
        error_message = str(error)

        logger.error(
            f"ThoughtSpot -> Power BI migration failed for job {job_id}: {error_message}",
            exc_info=True,
        )

        try:
            job_store.update_status(
                job_id=job_id,
                status=JobStatus.FAILED,
                error=error_message,
            )
        except TypeError:
            try:
                job_store.update_status(
                    job_id=job_id,
                    status=JobStatus.FAILED,
                )
            except Exception as update_error:
                logger.error(f"Failed to update failed job status: {update_error}")

        try:
            job_store.update_progress(
                job_id=job_id,
                percent=100,
                stage="failed",
                message=error_message,
                level="error",
            )
        except Exception:
            pass

        raise


def execute_discovery_job(job_id: str, file_paths: List[str]):
    """
    Backward-compatible old function name.
    """

    return execute_thoughtspot_powerbi_migration(
        job_id=job_id,
        file_paths=file_paths,
    )